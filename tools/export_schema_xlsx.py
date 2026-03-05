from __future__ import annotations

import argparse
import csv
import io
import shutil
from datetime import datetime
from pathlib import Path

import fitz
import pytesseract
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from PIL import Image
from pypdf import PdfReader


def _configure_tesseract() -> None:
    if shutil.which("tesseract"):
        return
    for candidate in (
        Path(r"C:\Program Files\Tesseract-OCR\tesseract.exe"),
        Path(r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe"),
    ):
        if candidate.exists():
            pytesseract.pytesseract.tesseract_cmd = str(candidate)
            return


def _ocr_image(image_bytes: bytes) -> str:
    image = Image.open(io.BytesIO(image_bytes))
    return (pytesseract.image_to_string(image, lang="eng") or "").strip()


def _ocr_pdf(pdf_path: Path) -> str:
    _configure_tesseract()
    doc = fitz.open(str(pdf_path))
    texts: list[str] = []
    try:
        for page in doc:
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
            texts.append(_ocr_image(pix.tobytes("png")))
    finally:
        doc.close()
    return "\n".join(filter(None, texts)).strip()


def _extract_text(pdf_path: Path) -> str:
    text = "\n".join((p.extract_text() or "") for p in PdfReader(str(pdf_path)).pages).strip()
    if len(text) >= 30:
        return text
    return _ocr_pdf(pdf_path)


def _fmt_date(value: str | None) -> str:
    if not value:
        return ""
    for fmt in ("%Y-%m-%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(value, fmt).strftime("%d.%m.%Y")
        except ValueError:
            continue
    return value


def _split_name(passenger_name: str | None, airline: str | None) -> tuple[str, str]:
    if not passenger_name:
        return "", ""
    parts = [p for p in passenger_name.split() if p]
    if len(parts) == 1:
        return parts[0], ""
    if (airline or "").lower() == "thy":
        return " ".join(parts[1:]), parts[0]
    return " ".join(parts[:-1]), parts[-1]


def _combine_dt(seg: dict, key_date: str, key_time: str) -> datetime | None:
    d = seg.get(key_date) or ""
    t = seg.get(key_time) or ""
    if not d or not t:
        return None
    try:
        return datetime.strptime(f"{d} {t}", "%Y-%m-%d %H:%M")
    except ValueError:
        return None


def _segment_block_type(segments: list[dict], seg: dict) -> str:
    role = seg.get("segment_role")
    if role == "gidis":
        return "gidiş"
    if role == "donus":
        return "dönüş"
    if role != "aktarma":
        return "segment"

    gidis_dt = None
    donus_dt = None
    for s in segments:
        if s.get("segment_role") == "gidis" and gidis_dt is None:
            gidis_dt = _combine_dt(s, "departure_date", "departure_time")
        if s.get("segment_role") == "donus" and donus_dt is None:
            donus_dt = _combine_dt(s, "departure_date", "departure_time")
    seg_dt = _combine_dt(seg, "departure_date", "departure_time")

    if donus_dt and seg_dt and seg_dt >= donus_dt:
        return "dönüş_aktarması"
    if gidis_dt and seg_dt and seg_dt <= gidis_dt:
        return "gidiş_aktarması"
    if donus_dt:
        return "dönüş_aktarması"
    return "gidiş_aktarması"


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--target-airport", default="AYT")
    parser.add_argument("--output", default="exports/final_export/schema_satir_satir_v2.xlsx")
    parser.add_argument("--output-csv", default="exports/final_export/schema_satir_satir_v2.csv")
    args = parser.parse_args()

    repo_root = Path(__file__).resolve().parent.parent
    exports_dir = repo_root / "exports"
    lists = [
        exports_dir / "sample_TK_10.txt",
        exports_dir / "sample_PC_10.txt",
        exports_dir / "sample_XQ_10.txt",
    ]
    files: list[Path] = []
    for list_path in lists:
        if not list_path.exists():
            continue
        for line in list_path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line:
                continue
            p = Path(line)
            if p.exists():
                files.append(p)

    seen = set()
    unique_files: list[Path] = []
    for f in files:
        key = str(f).lower()
        if key in seen:
            continue
        seen.add(key)
        unique_files.append(f)

    import sys

    sys.path.insert(0, str(repo_root / "backend"))
    from app.parser import parse_ticket_text  # pylint: disable=import-outside-toplevel

    columns = [
        "KAYNAK_DOSYA",
        "CINSIYET",
        "ISIM",
        "SOYISIM",
        "DUZENLENME_TARIHI",
        "BLOK",
        "UCUS_KODU",
        "ROTA",
        "KALKIS",
        "VARIS",
        "PICK-UP TIME",
        "PNR",
    ]
    rows: list[dict[str, str]] = []

    for pdf in unique_files:
        raw = _extract_text(pdf)
        parsed = parse_ticket_text(raw, target_airports=[args.target_airport]).get("parsed") or {}
        airline = (parsed.get("airline") or "").lower()
        isim, soyisim = _split_name(parsed.get("passenger_name"), airline)
        cinsiyet = parsed.get("cinsiyet") or "UNKNOW"
        pnr = parsed.get("pnr") or ""
        issue = _fmt_date(parsed.get("issue_date"))
        segments = parsed.get("segments") or []

        if not segments:
            dep_date = _fmt_date(parsed.get("date"))
            dep_time = parsed.get("time") or ""
            rows.append(
                {
                    "KAYNAK_DOSYA": pdf.name,
                    "CINSIYET": cinsiyet,
                    "ISIM": isim,
                    "SOYISIM": soyisim,
                    "DUZENLENME_TARIHI": issue,
                    "BLOK": "",
                    "UCUS_KODU": parsed.get("flight_no") or "",
                    "ROTA": f"{parsed.get('from') or ''}-{parsed.get('to') or ''}".strip("-"),
                    "KALKIS": f"{dep_date} {dep_time}".strip(),
                    "VARIS": "",
                    "PICK-UP TIME": "",
                    "PNR": pnr,
                }
            )
            continue

        for seg in segments:
            dep_date = _fmt_date(seg.get("departure_date"))
            arr_date = _fmt_date(seg.get("arrival_date"))
            dep_time = seg.get("departure_time") or ""
            arr_time = seg.get("arrival_time") or ""
            rows.append(
                {
                    "KAYNAK_DOSYA": pdf.name,
                    "CINSIYET": cinsiyet,
                    "ISIM": isim,
                    "SOYISIM": soyisim,
                    "DUZENLENME_TARIHI": issue,
                    "BLOK": _segment_block_type(segments, seg),
                    "UCUS_KODU": seg.get("flight_no") or "",
                    "ROTA": f"{seg.get('from') or ''}-{seg.get('to') or ''}".strip("-"),
                    "KALKIS": f"{dep_date} {dep_time}".strip(),
                    "VARIS": f"{arr_date} {arr_time}".strip(),
                    "PICK-UP TIME": "",
                    "PNR": pnr,
                }
            )

    out_xlsx = Path(args.output)
    if not out_xlsx.is_absolute():
        out_xlsx = repo_root / out_xlsx
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)

    out_csv = Path(args.output_csv)
    if not out_csv.is_absolute():
        out_csv = repo_root / out_csv
    out_csv.parent.mkdir(parents=True, exist_ok=True)

    with out_csv.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "schema"
    ws.append(columns)

    header_font = Font(bold=True)
    for col_idx in range(1, len(columns) + 1):
        ws.cell(row=1, column=col_idx).font = header_font

    cinsiyet_col_idx = columns.index("CINSIYET") + 1
    unknow_fill = PatternFill(fill_type="solid", start_color="FFEB9C", end_color="FFEB9C")
    unknow_font = Font(color="CC8000")
    for row in rows:
        ws.append([row.get(c, "") for c in columns])
        current_row = ws.max_row
        gender_cell = ws.cell(row=current_row, column=cinsiyet_col_idx)
        if str(gender_cell.value or "").strip().upper() == "UNKNOW":
            gender_cell.fill = unknow_fill
            gender_cell.font = unknow_font

    for col_idx, col_name in enumerate(columns, start=1):
        max_len = len(col_name)
        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            max_len = max(max_len, len(str(value or "")))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 80)

    wb.save(out_xlsx)
    print(f"rows={len(rows)}")
    print(f"xlsx={out_xlsx}")
    print(f"csv={out_csv}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
